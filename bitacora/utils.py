# sisboletas/bitacora/utils.py
import csv
import datetime
from django.http import HttpResponse
from django.utils.translation import gettext_lazy as _ 
from django.utils.html import strip_tags

# --- Imports para PDF ---
import io
from reportlab.lib.pagesizes import A4, landscape, letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import cm
from django.utils import timezone

# --- Imports para Excel ---
try:
    import openpyxl 
    from openpyxl.styles import Font, Alignment, Border, Side 
    OPENPYXL_AVAILABLE = True
    print("[DEBUG utils.py] openpyxl y componentes de estilo importados.")
except ImportError as e:
    OPENPYXL_AVAILABLE = False
    print(f"[ERROR utils.py] No se pudo importar openpyxl o sus componentes: {e}")
    GENERAR_EXCEL_IMPORT_ERROR = str(e) 
else:
    GENERAR_EXCEL_IMPORT_ERROR = None


def generar_respuesta_csv(queryset_logs, nombre_archivo_base="bitacora_export"):
    """
    Toma un queryset de LogEntry y devuelve una HttpResponse con el CSV.
    El queryset_logs ya debe venir filtrado desde la vista.
    """
    response = HttpResponse(
        content_type='text/csv',
        headers={'Content-Disposition': f'attachment; filename="{nombre_archivo_base}_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'},
    )

    writer = csv.writer(response, delimiter=';') 
    
    writer.writerow([
        _('Timestamp'), 
        _('Usuario (Nombre Completo)'), 
        _('Usuario (Username)'), 
        _('Acción'), 
        _('Tipo Recurso (App)'), 
        _('Tipo Recurso (Modelo)'),
        _('ID Objeto'), 
        _('Repr. Objeto'), 
        _('Detalles/Cambios'), 
        _('IP Remota')
    ])

    for log_entry in queryset_logs:
        actor_username = log_entry.actor.username if log_entry.actor else "Sistema"
        actor_fullname = log_entry.actor.get_full_name() if log_entry.actor and log_entry.actor.get_full_name() else actor_username       
        content_app = log_entry.content_type.app_label if log_entry.content_type else ""
        content_model = log_entry.content_type.model if log_entry.content_type else "" 
        timestamp_str = ""
        if log_entry.timestamp:
            try:
                timestamp_str = log_entry.timestamp.strftime("%Y-%m-%d %H:%M:%S %Z") 
                if not log_entry.timestamp.tzinfo:
                    timestamp_str = log_entry.timestamp.strftime("%Y-%m-%d %H:%M:%S")
            except AttributeError:
                timestamp_str = str(log_entry.timestamp)
        writer.writerow([
            timestamp_str,
            actor_fullname,
            actor_username,
            log_entry.get_action_display(), 
            content_app,
            content_model,
            log_entry.object_pk,
            log_entry.object_repr,
            log_entry.changes,
            log_entry.remote_addr
        ])
        
    return response



def generar_respuesta_pdf(queryset_logs, nombre_archivo_base="bitacora_export"):
    """
    Toma un queryset de LogEntry y devuelve una HttpResponse con el archivo PDF.
    """
    try:
        buffer = io.BytesIO()
        page_size = landscape(letter) 
        doc = SimpleDocTemplate(buffer, pagesize=page_size,
                                topMargin=1.5*cm, bottomMargin=1.5*cm,
                                leftMargin=1.5*cm, rightMargin=1.5*cm)
        
        styles = getSampleStyleSheet()
        cell_style = ParagraphStyle(
            'CellStyle',
            parent=styles['Normal'],
            fontSize=7,
            leading=9,
        )
        header_cell_style = ParagraphStyle(
            'HeaderCellStyle',
            parent=cell_style,
            fontName='Helvetica-Bold',
            alignment=1
        )

        elements = []

        title_style = styles['h2']
        title_style.alignment = 1
        title = Paragraph("Bitácora del Sistema", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.5*cm))
        data = [
            [
                Paragraph("Timestamp", header_cell_style), 
                Paragraph("Usuario", header_cell_style), 
                Paragraph("Acción", header_cell_style), 
                Paragraph("Recurso", header_cell_style), 
                Paragraph("ID Obj.", header_cell_style),
                Paragraph("Detalles", header_cell_style), 
                Paragraph("IP", header_cell_style)
            ]
        ]

        for log_entry in queryset_logs:
            actor_display = log_entry.actor.username if log_entry.actor else "Sistema"
            resource_repr = strip_tags(log_entry.object_repr)
            resource_display = (resource_repr[:40] + '...') if len(resource_repr) > 40 else resource_repr            
            changes_clean = strip_tags(log_entry.changes)
            changes_display = (changes_clean[:60] + '...') if len(changes_clean) > 60 else changes_clean            
            ts_display = log_entry.timestamp.strftime("%d/%m/%y %H:%M") if log_entry.timestamp else ""

            data.append([
                Paragraph(ts_display, cell_style),
                Paragraph(actor_display, cell_style),
                Paragraph(log_entry.get_action_display(), cell_style),
                Paragraph(resource_display, cell_style),
                Paragraph(str(log_entry.object_pk or ""), cell_style),
                Paragraph(changes_display, cell_style),
                Paragraph(log_entry.remote_addr or "-", cell_style)
            ])

        if len(data) == 1:
            elements.append(Paragraph("No hay registros para mostrar con los filtros actuales.", styles['Normal']))
        else:
            available_width = doc.width 
            col_widths = [
                available_width * 0.15, # Timestamp
                available_width * 0.15, # Usuario
                available_width * 0.10, # Acción
                available_width * 0.20, # Recurso
                available_width * 0.08, # ID Obj
                available_width * 0.22, # Detalles
                available_width * 0.10  # IP
            ]
            
            table = Table(data, colWidths=col_widths, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#40516F")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'), # Alinear a la izquierda por defecto
                ('ALIGN', (0, 0), (0, -1), 'CENTER'), # Timestamp centrado
                ('ALIGN', (2, 0), (2, -1), 'CENTER'), # Acción centrada
                ('ALIGN', (4, 0), (4, -1), 'CENTER'), # ID Obj centrado
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8), # Tamaño de fuente para encabezado
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#E6E6E6")), # Color claro para filas de datos
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ]))
            elements.append(table)

        doc.build(elements)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer, 
            content_type='application/pdf',
            headers={'Content-Disposition': f'attachment; filename="{nombre_archivo_base}_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'}
        )
        return response
    except ImportError:
        return HttpResponse("La librería 'reportlab' es necesaria para exportar a PDF. Por favor, instálala.", status=501)
    except Exception as e:
        print(f"Error generando PDF de bitácora: {e}")
        return HttpResponse(f"Error al generar el PDF: {e}", status=500)
    

def generar_respuesta_excel(queryset_logs, nombre_archivo_base="bitacora_export"):
    """
    Toma un queryset de LogEntry y devuelve una HttpResponse con el archivo Excel (.xlsx).
    """
    if not OPENPYXL_AVAILABLE:
        error_message = "La librería 'openpyxl' es necesaria para exportar a Excel. Por favor, instálala (pip install openpyxl)."
        if GENERAR_EXCEL_IMPORT_ERROR:
            error_message += f" Detalle del error de importación: {GENERAR_EXCEL_IMPORT_ERROR}"
        return HttpResponse(error_message, status=501)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "BitacoraLogs"

    # --- Estilos ---
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFFFF') # Texto blanco
    header_fill = openpyxl.styles.PatternFill(start_color='40516F', end_color='40516F', fill_type='solid') # Azul oscuro
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True) # wrap_text para celdas con mucho texto
    
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    # --- Encabezados ---
    headers = [
        'Timestamp', 'Usuario (Nombre Completo)', 'Usuario (Username)', 'Acción', 
        'Tipo Recurso (App)', 'Tipo Recurso (Modelo)',
        'ID Objeto', 'Repr. Objeto', 'Detalles/Cambios', 'IP Remota'
    ]
    sheet.append(headers)

    for col_num, header_title in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border

    # --- Datos ---
    row_num = 2 # Empezar desde la fila 2 para los datos
    for log_entry in queryset_logs:
        actor_username = log_entry.actor.username if log_entry.actor else "Sistema"
        actor_fullname = log_entry.actor.get_full_name() if log_entry.actor and log_entry.actor.get_full_name() else actor_username
        content_app = log_entry.content_type.app_label if log_entry.content_type else ""
        content_model = log_entry.content_type.model if log_entry.content_type else ""       
        timestamp_excel = None
        if log_entry.timestamp:
            if log_entry.timestamp.tzinfo:
                timestamp_excel = log_entry.timestamp.astimezone(None)
            else:
                timestamp_excel = log_entry.timestamp
        
        changes_cleaned = strip_tags(log_entry.changes)

        # --- MANEJO DEL TIMESTAMP (MÁS ROBUSTO) ---
        timestamp_excel = None
        if log_entry.timestamp:
            local_dt = timezone.localtime(log_entry.timestamp) # Convierte a TIME_ZONE si es aware UTC, o localiza si es naive
            timestamp_excel = local_dt.replace(tzinfo=None)    # Hacerlo naive
        data_row = [
            timestamp_excel,
            actor_fullname,
            actor_username,
            log_entry.get_action_display(),
            content_app,
            content_model,
            log_entry.object_pk,
            log_entry.object_repr,
            changes_cleaned,
            log_entry.remote_addr
        ]
        sheet.append(data_row)

        # Aplicar estilo a las celdas de datos
        for col_num in range(1, len(data_row) + 1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.alignment = left_alignment
            cell.border = thin_border
            if col_num == 1 and timestamp_excel: # Formato de fecha para la columna Timestamp
                 cell.number_format = 'DD/MM/YYYY HH:MM:SS'
        row_num += 1

    # --- Ajustar Ancho de Columnas dinamico ---
    column_widths = {'A': 22, 'B': 25, 'C': 20, 'D': 15, 'E': 20, 'F': 20, 'G': 12, 'H': 40, 'I': 60, 'J': 18}
    for col_letter, width in column_widths.items():
        sheet.column_dimensions[col_letter].width = width   
    excel_buffer = io.BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)
    
    response = HttpResponse(
        excel_buffer.getvalue(), 
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    timestamp_export = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo_base}_{timestamp_export}.xlsx"'
    return response