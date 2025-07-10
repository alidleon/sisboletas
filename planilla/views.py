import logging
from django.shortcuts import render, redirect, get_object_or_404, HttpResponse
from django.contrib.auth.decorators import login_required, permission_required
import os
from django.conf import settings
from django.urls import reverse 
from urllib.parse import urlencode 
from datetime import date
from django.contrib import messages
from .forms import DetalleBonoTeForm
from .forms import PlanillaForm 
from datetime import datetime
from django.core.exceptions import ValidationError
from decimal import Decimal, InvalidOperation
from django.utils import timezone 
from collections import defaultdict 
from operator import attrgetter 
# Imports para ReportLab
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib import colors
from reportlab.lib.units import inch, cm
from reportlab.lib.utils import ImageReader
import os
from django.conf import settings
import io
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger 
from calendar import month_name
from openpyxl.drawing.image import Image
from django.db import transaction, IntegrityError 
from django.db.models import Q  
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger 
# --- Importaciones de Modelos Originales ---
from .models import (
    Planilla,
    DetalleBonoTe,
    PrincipalDesignacionExterno,
    PrincipalPersonalExterno,
    PrincipalCargoExterno,
    PrincipalUnidadExterna,
    PrincipalSecretariaExterna,
)
from .utils import get_processed_planilla_details, generar_pdf_bonote_detalles, generar_pdf_lista_planillas
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.drawing.image import Image as OpenpyxlImage
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    class Font: pass
    class Alignment: pass
    class PatternFill: pass
    class Border: pass
    class Side: pass
    class Image: pass

logger = logging.getLogger(__name__)

from .forms import DetalleBonoTeForm, PlanillaForm, EditarPlanillaForm

try:
    from reportes.models import PlanillaAsistencia, DetalleAsistencia
    REPORTES_APP_AVAILABLE = True
except ImportError:
    PlanillaAsistencia = None
    DetalleAsistencia = None
    REPORTES_APP_AVAILABLE = False
    logging.error("ERROR CRÍTICO: No se pueden importar modelos de la app 'reportes'.")


# ---------------------------------

@login_required
def seleccionar_tipo_planilla(request):
    tipos_disponibles = Planilla.TIPO_CHOICES 
    if request.method == 'POST':
        tipo = request.POST.get('tipo')
        tipos_validos_keys = dict(tipos_disponibles).keys()
        if tipo in tipos_validos_keys:
            return redirect('crear_planilla', tipo=tipo) 
        else:
            messages.error(request, 'Seleccione un tipo de planilla válido.')
            return render(request, 'planillas/seleccionar_tipo_planilla.html', {'tipos_planilla': tipos_disponibles})
    else: 
        context = {
            'tipos_planilla': tipos_disponibles,
        }
        return render(request, 'planillas/seleccionar_tipo_planilla.html', context)
#-------------------------------------------------------------------------

EXTERNAL_TYPE_MAP = {
        'planta': 'ASEGURADO',      
        'contrato': 'CONTRATO',
        'consultor': 'CONSULTOR EN LINEA',
    }

logger = logging.getLogger(__name__)


@login_required
@permission_required('planilla.add_planilla', raise_exception=True)
def crear_planilla_bono_te(request):
    if not REPORTES_APP_AVAILABLE:
        messages.error(request, "Error crítico: La funcionalidad de reportes de asistencia no está disponible.")
        return redirect('lista_planillas')
    tipo_filtro_seleccionado = request.GET.get('tipo_filtro', None)

    if request.method == 'POST':
        form = PlanillaForm(request.POST, tipo_filtro=tipo_filtro_seleccionado) 
        if form.is_valid():
            selected_pa_base = form.cleaned_data['planilla_asistencia_base_selector']
            dias_habiles_ingresados = form.cleaned_data['dias_habiles']
            
            mes_derivado = selected_pa_base.mes
            anio_derivado = selected_pa_base.anio
            tipo_planilla_derivado = selected_pa_base.tipo 
            if tipo_filtro_seleccionado and tipo_planilla_derivado != tipo_filtro_seleccionado:
                messages.error(request, "Error: El tipo de la planilla de asistencia seleccionada no coincide con el tipo filtrado.")
                form_con_error = PlanillaForm(request.POST, tipo_filtro=tipo_filtro_seleccionado) 
                context_error = {
                    'planilla_form': form_con_error,
                    'tipos_planilla_choices': Planilla.TIPO_CHOICES, 
                    'tipo_filtro_actual': tipo_filtro_seleccionado,
                    'titulo_vista': f"Crear Planilla Bono TE (Filtrado por: {dict(Planilla.TIPO_CHOICES).get(tipo_filtro_seleccionado,'Todos') if tipo_filtro_seleccionado else 'Todos los Tipos'})"
                }
                return render(request, 'planillas/crear_planilla.html', context_error)
            
            plan_asistencia_validada = selected_pa_base
            detalles_asistencia_dict = {}
            try:
                detalles_asist_qs = DetalleAsistencia.objects.filter(planilla_asistencia=plan_asistencia_validada)
                detalles_asistencia_dict = {
                    da.personal_externo_id: da for da in detalles_asist_qs if da.personal_externo_id is not None
                }
                logger.info(f"Cargados {len(detalles_asistencia_dict)} detalles de asistencia desde la PlanillaAsistencia ID {plan_asistencia_validada.id}.")
            except Exception as e_asist:
                logger.error(f"Error inesperado cargando detalles de asistencia de PA ID {plan_asistencia_validada.id}: {e_asist}", exc_info=True)
                messages.error(request, f"Ocurrió un error al obtener los detalles de la asistencia seleccionada: {e_asist}")
                context_error = {'planilla_form': PlanillaForm(request.POST, tipo_filtro=tipo_filtro_seleccionado), 'titulo_vista': "Crear Planilla Bono TE", 'tipos_planilla_choices': Planilla.TIPO_CHOICES, 'tipo_filtro_actual': tipo_filtro_seleccionado}
                return render(request, 'planillas/crear_planilla.html', context_error)

            try:
                with transaction.atomic(using='default'):
                    planilla_bonote = Planilla(
                        planilla_asistencia_base=plan_asistencia_validada,
                        mes=mes_derivado,
                        anio=anio_derivado,
                        tipo=tipo_planilla_derivado,
                        dias_habiles=dias_habiles_ingresados,
                        estado='pendiente',
                        usuario_elaboracion=request.user,
                        fecha_elaboracion=timezone.now().date()
                    )
                    planilla_bonote.full_clean()
                    planilla_bonote.save()
                    logger.info(f"Creada Planilla (Bono TE) ID {planilla_bonote.id} para PA Base ID {plan_asistencia_validada.id}.")

                    detalles_bonote_a_crear = []
                    if not detalles_asistencia_dict:
                        messages.warning(request, f"Planilla Bono TE creada (ID: {planilla_bonote.id}), pero la Planilla de Asistencia base seleccionada no contenía detalles de personal.")
                    else:
                        logger.info(f"Preparando {len(detalles_asistencia_dict)} registros DetalleBonoTe desde la asistencia base...")
                        for personal_id_asistencia, detalle_asistencia_origen in detalles_asistencia_dict.items():
                            detalle_bt = DetalleBonoTe(
                                id_planilla=planilla_bonote,
                                personal_externo_id=personal_id_asistencia,
                                mes=planilla_bonote.mes,
                                faltas=getattr(detalle_asistencia_origen, 'faltas_dias', 0),
                                vacacion=getattr(detalle_asistencia_origen, 'vacacion', 0),
                                viajes=getattr(detalle_asistencia_origen, 'viajes', 0),
                                bajas_medicas=getattr(detalle_asistencia_origen, 'bajas_medicas', 0),
                                pcgh=getattr(detalle_asistencia_origen, 'pcgh', 0),
                                psgh=getattr(detalle_asistencia_origen, 'psgh', 0),
                                perm_excep=getattr(detalle_asistencia_origen, 'perm_excep', 0),
                                asuetos=getattr(detalle_asistencia_origen, 'asuetos', 0),
                                pcgh_embar_enf_base=getattr(detalle_asistencia_origen, 'pcgh_embar_enf_base', 0),
                                
                                observaciones_asistencia=getattr(detalle_asistencia_origen, 'observaciones', ''),
                            )
                            detalles_bonote_a_crear.append(detalle_bt)

                        if detalles_bonote_a_crear:
                            DetalleBonoTe.objects.bulk_create(detalles_bonote_a_crear)
                            logger.info(f"Creados {len(detalles_bonote_a_crear)} registros DetalleBonoTe para Planilla {planilla_bonote.id}.")
                            detalles_creados_para_recalculo = DetalleBonoTe.objects.filter(id_planilla=planilla_bonote)
                            detalles_a_actualizar_con_calculos = []
                            for detalle_bt_creado in detalles_creados_para_recalculo:
                                detalle_bt_creado.calcular_valores()
                                detalles_a_actualizar_con_calculos.append(detalle_bt_creado)
                            campos_calculados = ['dias_no_pagados', 'dias_pagados', 'total_ganado', 'liquido_pagable']
                            DetalleBonoTe.objects.bulk_update(detalles_a_actualizar_con_calculos, campos_calculados)
                            logger.info(f"Valores recalculados y actualizados para {len(detalles_a_actualizar_con_calculos)} detalles Bono TE.")
                            messages.success(request, f"Planilla Bono TE ({planilla_bonote.get_tipo_display()} {planilla_bonote.mes}/{planilla_bonote.anio}) creada exitosamente con {len(detalles_bonote_a_crear)} registros, basada en la asistencia seleccionada.")
                        else:
                             messages.warning(request, f"Planilla Bono TE creada, pero no se generaron detalles.")
                return redirect('lista_planillas')
            except ValidationError as e_model_validation:
                logger.warning(f"Error de validación del modelo al crear Planilla Bono TE: {e_model_validation.message_dict}")
                for field, errors in e_model_validation.message_dict.items():
                    form_field_name = 'planilla_asistencia_base_selector' if field == 'planilla_asistencia_base' else field
                    for error_msg in errors:
                        form.add_error(form_field_name, error_msg) 
            except IntegrityError as e_db_integrity:
                logger.error(f"Error de integridad en la BD al crear Planilla/Detalles Bono TE: {e_db_integrity}", exc_info=True)
                form.add_error(None, f"Error de base de datos. ({e_db_integrity})")
            except Exception as e_general_processing:
                 logger.error(f"Error general procesando la creación: {e_general_processing}", exc_info=True)
                 form.add_error(None, f"Error inesperado: {e_general_processing}")
    else: 
        form = PlanillaForm(tipo_filtro=tipo_filtro_seleccionado)
    context = {
        'planilla_form': form,
        'tipos_planilla_choices': Planilla.TIPO_CHOICES, 
        'tipo_filtro_actual': tipo_filtro_seleccionado,  
        'titulo_vista': f"Crear Planilla Bono TE (Filtrando por: {dict(Planilla.TIPO_CHOICES).get(tipo_filtro_seleccionado,'Todos los Tipos') if tipo_filtro_seleccionado else 'Todos los Tipos'})"
    }
    return render(request, 'planillas/crear_planilla.html', context)

#------------------------------------------
@login_required
@permission_required('planilla.view_planilla', raise_exception=True)
def lista_planillas(request):
    """
    Muestra una lista PAGINADA y FILTRABLE de todas las Planillas de Bono TE.
    """
    logger.debug(f"Vista lista_planillas (Bono TE) llamada. GET params: {request.GET.urlencode()}")
    queryset = Planilla.objects.all().order_by('-anio', '-mes', 'tipo')
    filtro_anio = request.GET.get('anio', '').strip()
    filtro_mes = request.GET.get('mes', '').strip()
    filtro_tipo = request.GET.get('tipo', '').strip()
    filtro_estado = request.GET.get('estado', '').strip()
    if filtro_anio:
        try:
            queryset = queryset.filter(anio=int(filtro_anio))
        except (ValueError, TypeError):
            pass
    if filtro_mes:
        try:
            queryset = queryset.filter(mes=int(filtro_mes))
        except (ValueError, TypeError):
            pass

    if filtro_tipo:
        queryset = queryset.filter(tipo=filtro_tipo)

    if filtro_estado:
        queryset = queryset.filter(estado=filtro_estado)
    items_por_pagina = 10
    paginator = Paginator(queryset, items_por_pagina)
    page_number_str = request.GET.get('page', '1')
    
    try:
        page_obj = paginator.page(page_number_str)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
        
    logger.info(f"Mostrando página {page_obj.number} de {paginator.num_pages} para Planillas Bono TE (Total filtrado: {paginator.count}).")    
    querystring = request.GET.copy()
    if 'page' in querystring:
        del querystring['page']        
    context = {
        'page_obj': page_obj,
        'titulo_vista': "Lista de Planillas Bono Te ",
        'valores_filtro': request.GET,
        'querystring': querystring.urlencode(),
        'tipos_disponibles': Planilla.TIPO_CHOICES,
        'estados_disponibles': Planilla.ESTADO_CHOICES,
    }
    
    return render(request, 'planillas/lista_planillas.html', context)

#---------------------------------------
@login_required
@permission_required('planilla.view_detallebonote', raise_exception=True)
def lista_bono_te(request):
    detalles_bono_te = DetalleBonoTe.objects.select_related('id_planilla').all()
    return render(request, 'planillas/lista_bono_te.html', {'detalles_bono_te': detalles_bono_te})
#----------------------------------------------------------

@login_required
@permission_required('planilla.change_detallebonote', raise_exception=True)
def editar_bono_te(request, detalle_id):
    detalle_bono_te = get_object_or_404(
        DetalleBonoTe.objects.select_related('id_planilla'), 
        pk=detalle_id
    )
    planilla = detalle_bono_te.id_planilla
    dias_habiles_planilla = planilla.dias_habiles if planilla else None
    persona_externa = None
    item_externo = 'N/A'
    cargo_externo = 'N/A'
    personal_externo_id = detalle_bono_te.personal_externo_id
    if personal_externo_id:
        try:
            persona_externa = PrincipalPersonalExterno.objects.using('personas_db').get(pk=personal_externo_id)
        except PrincipalPersonalExterno.DoesNotExist:
            logger.warning(f"No se encontró PrincipalPersonalExterno ID {personal_externo_id} en 'personas_db'")
        except Exception as e_pers:
            logger.error(f"Error consultando PrincipalPersonalExterno ID {personal_externo_id}: {e_pers}", exc_info=True)
        try:
            designacion = PrincipalDesignacionExterno.objects.using('personas_db') \
                .select_related('cargo') \
                .filter(personal_id=personal_externo_id, estado='ACTIVO') \
                .order_by('-id').first() 

            if designacion:
                item_externo = designacion.item if designacion.item is not None else 'N/A'
                cargo_externo = designacion.cargo.nombre_cargo if designacion.cargo else 'N/A'
            else:
                 logger.warning(f"No se encontró designación externa ACTIVA para Persona ID {personal_externo_id} en 'personas_db'.")
        except Exception as e_desig:
            logger.error(f"Error consultando PrincipalDesignacionExterno para Persona ID {personal_externo_id}: {e_desig}", exc_info=True)

    if request.method == 'POST':
        form = DetalleBonoTeForm(request.POST, instance=detalle_bono_te) 
        if form.is_valid():
            form.save()
            messages.success(request, 'Detalle Bono TE editado correctamente.')
            redirect_secretaria = request.POST.get('redirect_secretaria', '')
            redirect_unidad = request.POST.get('redirect_unidad', '')
            redirect_q = request.POST.get('redirect_q', '')
            base_url = reverse('ver_detalles_bono_te', kwargs={'planilla_id': planilla.id})
            params = {}
            if redirect_secretaria: params['secretaria'] = redirect_secretaria
            if redirect_unidad: params['unidad'] = redirect_unidad
            if redirect_q: params['q'] = redirect_q
            if params: params['buscar'] = 'true'
            redirect_url = f"{base_url}?{urlencode(params)}" if params else base_url
            return redirect(redirect_url)
        else:
            messages.error(request, 'Por favor, corrige los errores en el formulario.')
    else: 
        form = DetalleBonoTeForm(
            instance=detalle_bono_te,
            initial={'dias_habiles': dias_habiles_planilla} 
        )

    context = {
        'form': form,
        'detalle_bono_te': detalle_bono_te, 
        'dias_habiles': dias_habiles_planilla,
        'persona_externa': persona_externa, 
        'item_externo': item_externo,
        'cargo_externo': cargo_externo,
        'redirect_secretaria': request.GET.get('secretaria', ''),
        'redirect_unidad': request.GET.get('unidad', ''),
        'redirect_q': request.GET.get('q', ''),
    }
    return render(request, 'planillas/editar_bono_te.html', context)

#---------------------------------------
@login_required
@permission_required('planilla.delete_detallebonote', raise_exception=True)
def borrar_bono_te(request, detalle_id):
    detalle_bono_te = get_object_or_404(DetalleBonoTe, pk=detalle_id)
    planilla_id = detalle_bono_te.id_planilla_id
    persona_nombre = f"ID Externo {detalle_bono_te.personal_externo_id}"
    try:
         if detalle_bono_te.personal_externo_id:
              persona = PrincipalPersonalExterno.objects.using('personas_db').get(pk=detalle_bono_te.personal_externo_id)
              persona_nombre = persona.nombre_completo or persona_nombre
    except:
        pass 

    if request.method == 'POST':
        detalle_bono_te.delete()
        messages.success(request, f'Detalle Bono TE para {persona_nombre} borrado correctamente.')
        return redirect('ver_detalles_bono_te', planilla_id=planilla_id)

    return render(request, 'planillas/borrar_bono_te.html', {
        'detalle_bono_te': detalle_bono_te,
        'persona_nombre': persona_nombre 
        })
#-------------------------------------

@login_required
@permission_required('planilla.change_planilla', raise_exception=True)
def editar_planilla(request, planilla_id):
    planilla_instancia = get_object_or_404(Planilla, pk=planilla_id)
    if request.method == 'POST':
        form = EditarPlanillaForm(request.POST, instance=planilla_instancia)
        if form.is_valid():
            try:
                planilla_editada = form.save(commit=False)
                if 'dias_habiles' in form.changed_data:
                    logger.info(f"Días hábiles cambiados para Planilla ID {planilla_editada.id}. Se recalcularán los detalles.")
                    planilla_editada.save()
                    detalles_a_recalcular = DetalleBonoTe.objects.filter(id_planilla=planilla_editada)
                    detalles_a_actualizar_bulk = []
                    for detalle in detalles_a_recalcular:
                        detalle.calcular_valores()
                        detalles_a_actualizar_bulk.append(detalle)
                    if detalles_a_actualizar_bulk:
                        campos_calculados = ['dias_no_pagados', 'dias_pagados', 'total_ganado', 'liquido_pagable']
                        DetalleBonoTe.objects.bulk_update(detalles_a_actualizar_bulk, campos_calculados)
                        logger.info(f"Recalculados {len(detalles_a_actualizar_bulk)} detalles para Planilla ID {planilla_editada.id}.")
                    form.save_m2m() 
                else:
                    planilla_editada.save() 
                    form.save_m2m()
                messages.success(request, f'Planilla "{planilla_editada}" actualizada correctamente.')
                return redirect('lista_planillas')
            except Exception as e:
                logger.error(f"Error al guardar la planilla editada ID {planilla_instancia.id}: {e}", exc_info=True)
                messages.error(request, f"Ocurrió un error al guardar los cambios: {e}")
        else:
            messages.error(request, 'Por favor, corrige los errores en el formulario.')
    else: 
        form = EditarPlanillaForm(instance=planilla_instancia)
    context = {
        'form': form,
        'planilla': planilla_instancia, 
        'titulo_vista': f"Editar Planilla Bono TE: {planilla_instancia.mes}/{planilla_instancia.anio} ({planilla_instancia.get_tipo_display()})"
    }
    return render(request, 'planillas/editar_planilla.html', context)

#----------------------------------------------
@login_required
@permission_required('planilla.delete_planilla', raise_exception=True)
def borrar_planilla(request, planilla_id):
    planilla = get_object_or_404(Planilla, pk=planilla_id)
    if request.method == 'POST':
        planilla.delete()
        messages.success(request, 'Planilla borrada correctamente.')
        return redirect('lista_planillas')
    return render(request, 'planillas/borrar_planilla.html', {'planilla': planilla})

#------------------------------------------------
@login_required
@permission_required('planilla.view_detallebonote', raise_exception=True) 
def ver_detalles_bono_te(request, planilla_id):
    logger.debug(f"Vista ver_detalles_bono_te llamada para planilla_id={planilla_id}. GET: {request.GET.urlencode()}")
    try:
        items_por_pagina_vista = 20 
        processed_data = get_processed_planilla_details(request, planilla_id, items_por_pagina=items_por_pagina_vista)

        planilla_obj_vista = processed_data.get('planilla') 
        if not planilla_obj_vista and processed_data.get('error_message'):
            messages.error(request, processed_data['error_message'])
            return redirect('lista_planillas')
        if not planilla_obj_vista: 
            messages.error(request, "Planilla de Bono TE no encontrada.")
            return redirect('lista_planillas')
        if processed_data.get('error_message') and planilla_obj_vista:
             messages.warning(request, processed_data['error_message'])

        page_obj_actual = processed_data.get('page_obj')

        context = {
            'planilla': planilla_obj_vista, 
            'all_secretarias': processed_data.get('all_secretarias', []),
            'unidades_for_select': processed_data.get('unidades_for_select', []),
            'selected_secretaria_id': processed_data.get('selected_secretaria_id'),
            'selected_unidad_id': processed_data.get('selected_unidad_id'),
            'search_term': processed_data.get('search_term', ''),
            'search_active': processed_data.get('search_active', False),
            'page_obj': page_obj_actual, 
            'titulo_vista': f"Detalles Bono TE - {planilla_obj_vista.get_tipo_display()} {planilla_obj_vista.mes}/{planilla_obj_vista.anio}"
        }
        
        if page_obj_actual:
            logger.debug(f"VISTA Planilla: Renderizando con page_obj: Pág {page_obj_actual.number}/{page_obj_actual.paginator.num_pages}, Total items: {page_obj_actual.paginator.count}")
        else:
            logger.warning("VISTA Planilla: Renderizando SIN page_obj (o es None).")
            
        return render(request, 'planillas/ver_detalles_bono_te.html', context)

    except Exception as e_view:
        logger.error(f"Error inesperado en vista ver_detalles_bono_te ID {planilla_id}: {e_view}", exc_info=True)
        messages.error(request, "Ocurrió un error inesperado al mostrar los detalles del bono.")
        return redirect('lista_planillas')




MESES_ESPANOL = [
    None, 
    "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO",
    "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"
]
def safe_decimal(value, default=Decimal('0')):
    if value is None: return default
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError, TypeError):
        return default

def safe_int(value, default=0):
    if value is None: return default
    try:
        return int(Decimal(str(value).strip()))
    except (InvalidOperation, ValueError, TypeError):
        return default
#----------------------------------

@login_required
@permission_required('planilla.view_planilla', raise_exception=True) 
def exportar_planilla_xlsx(request, planilla_id):
    if not OPENPYXL_AVAILABLE:
        messages.error(request, "La funcionalidad de exportación a Excel no está disponible (falta 'openpyxl').")
        return redirect('ver_detalles_bono_te', planilla_id=planilla_id)

    logger.info(f"Solicitud de exportación XLSX para Planilla ID {planilla_id} (Agrupado por Unidad)")

    try:
        processed_data = get_processed_planilla_details(request, planilla_id, return_all_for_export=True)
    except Exception as e_get_data:
        logger.error(f"Error crítico obteniendo datos para exportar Planilla {planilla_id}: {e_get_data}", exc_info=True)
        messages.error(request, f"No se pudieron obtener los datos para exportar: {e_get_data}")
        return redirect('lista_planillas') 

    planilla = processed_data.get('planilla')
    detalles_agrupados_export = processed_data.get('detalles_agrupados_por_unidad_export', {})
    orden_unidades_export = processed_data.get('orden_unidades_export', [])
    
    selected_secretaria_id = processed_data.get('selected_secretaria_id')
    selected_unidad_id = processed_data.get('selected_unidad_id') 
    search_term = processed_data.get('search_term', '')


    if not planilla:
         messages.error(request, "No se pudo encontrar la información de la planilla para exportar.")
         return redirect('lista_planillas')
    if processed_data.get('error_message') and not detalles_agrupados_export:
        messages.error(request, f"Error al preparar datos para exportar: {processed_data['error_message']}")
        return redirect('ver_detalles_bono_te', planilla_id=planilla_id)
    total_detalles_a_exportar = sum(len(lista) for lista in detalles_agrupados_export.values())

    if not detalles_agrupados_export:
        msg_no_detalles = (f"No se encontraron detalles para la Planilla ID {planilla_id} "
                           f"con los filtros aplicados. Se generará un archivo Excel solo con la cabecera.")
        logger.warning(msg_no_detalles)
        messages.warning(request, msg_no_detalles)
    else:
        logger.info(f"Se procesarán {len(orden_unidades_export)} grupos de unidades para la planilla ID {planilla_id}. Total detalles: {total_detalles_a_exportar}")

    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = f"BonoTE_{planilla.mes}_{planilla.anio}"[:31] 
        inst_header_bold_font = Font(name='Calibri', size=9, bold=True)
        title_font = Font(name='Calibri', size=14, bold=True) 
        subtitle_font = Font(name='Calibri', size=11, bold=True) 
        value_font = Font(name='Calibri', size=9)
        header_font = Font(name='Calibri', size=9, bold=True) 
        unidad_header_font = Font(name='Calibri', size=10, bold=True)
        wrap_left_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        centered_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        right_alignment = Alignment(horizontal='right', vertical='center')
        decimal_format = '#,##0.00'; integer_format = '#,##0'
        header_fill = PatternFill(start_color='00FFFF', end_color='00FFFF', fill_type='solid') 
        thin_border_side = Side(border_style="thin", color="FF000000") 
        header_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
        data_headers = [
            'Nro.', 'Item', 'CI', 'Nombre Completo', 'Cargo', 
            'Días Háb.', 'Faltas', 'Vacac.', 'Viajes', 'B.Médicas', 'PCGH', 'PSGH', 'P.Excep',
            'Asuetos', 'PCGH Emb/Enf', 'D.No Pag.', 'D. Pag.', 'Total Ganado',
            'Desc.', 'Líquido Pagable'
        ]
        num_data_columns = len(data_headers)
        current_row = 1 
        institucional_text = (
            "GOBIERNO AUTONOMO DEPARTAMENTAL DE POTOSI\n"
            "SECRETARIA DEPTAL. ADMINISTRATIVA FINANCIERA\n"
            "UNIDAD DE RECURSOS HUMANOS"
        )
        sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row + 2, end_column=max(5, num_data_columns // 3)) # Fusionar para texto institucional
        inst_cell = sheet.cell(row=current_row, column=1, value=institucional_text)
        inst_cell.font = inst_header_bold_font
        inst_cell.alignment = wrap_left_alignment        
        image_path = None
        try:
            logo_filename = 'gadp.png'
            if settings.STATICFILES_DIRS:
                potential_path = os.path.join(settings.STATICFILES_DIRS[0], 'img', logo_filename)
                if os.path.exists(potential_path): image_path = potential_path
            if not image_path:
                potential_path_base = os.path.join(settings.BASE_DIR, 'static', 'img', logo_filename)
                if os.path.exists(potential_path_base): image_path = potential_path_base
            
            if image_path:
                img = OpenpyxlImage(image_path) 
                img.height = 75 
                img.width = 65  
                logo_anchor_column_letter = get_column_letter(num_data_columns -1 if num_data_columns > 2 else num_data_columns) 
                logo_anchor_cell = f'{logo_anchor_column_letter}{current_row}'
                sheet.add_image(img, logo_anchor_cell)
                logger.info(f"Logo añadido al Excel en la celda: {logo_anchor_cell} desde {image_path}")
            else:
                logger.warning(f"Logo '{logo_filename}' no encontrado en rutas probadas para Excel.")
        except Exception as e_img:
             logger.error(f"Error al procesar o añadir la imagen al Excel: {e_img}", exc_info=True)
        current_row += 3 

        # --- Título Principal de la Planilla ---
        sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=num_data_columns)
        title_text = f"PLANILLA DE PAGO BONO DE TÉ - PERSONAL {planilla.get_tipo_display().upper()}"
        title_cell = sheet.cell(row=current_row, column=1, value=title_text)
        title_cell.font = title_font
        title_cell.alignment = centered_alignment
        current_row += 1

        # --- Período (Mes y Año) ---
        sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=num_data_columns)
        nombre_mes_str = f"MES {planilla.mes}"
        try:
            mes_numero = int(planilla.mes)
            if 1 <= mes_numero <= 12: nombre_mes_str = MESES_ESPANOL[mes_numero]
        except (IndexError, TypeError, ValueError): pass 
        periodo_text = f"CORRESPONDIENTE AL MES DE {nombre_mes_str} DE {planilla.anio}"
        periodo_cell = sheet.cell(row=current_row, column=1, value=periodo_text)
        periodo_cell.font = subtitle_font
        periodo_cell.alignment = centered_alignment
        current_row += 1
        filter_info_parts = []
        if selected_secretaria_id:
            try:
                sec = PrincipalSecretariaExterna.objects.using('personas_db').get(pk=selected_secretaria_id)
                filter_info_parts.append(f"Secretaría: {sec.nombre_secretaria}")
            except: filter_info_parts.append(f"Sec.ID: {selected_secretaria_id}")
        if selected_unidad_id: 
            try:
                uni = PrincipalUnidadExterna.objects.using('personas_db').get(pk=selected_unidad_id)
                filter_info_parts.append(f"Unidad: {uni.nombre_unidad}")
            except: filter_info_parts.append(f"Uni.ID: {selected_unidad_id}")
        if search_term:
             filter_info_parts.append(f"Busq: '{search_term}'")

        if filter_info_parts:
            sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=num_data_columns)
            filter_text = "Filtros Aplicados: " + " | ".join(filter_info_parts)
            filter_cell = sheet.cell(row=current_row, column=1, value=filter_text)
            filter_cell.font = Font(name='Calibri', size=8, italic=True)
            filter_cell.alignment = left_alignment
            current_row += 1
        current_row += 1 
        # --- Iterar sobre Unidades y sus Detalles ---
        nro_general_item_excel = 0
        for nombre_unidad_actual_excel in orden_unidades_export:
            detalles_de_esta_unidad_excel = detalles_agrupados_export.get(nombre_unidad_actual_excel, [])
            if not detalles_de_esta_unidad_excel:
                continue
            sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=num_data_columns)
            unidad_text_excel = f"UNIDAD: {nombre_unidad_actual_excel.upper()}"
            unidad_cell_excel = sheet.cell(row=current_row, column=1, value=unidad_text_excel)
            unidad_cell_excel.font = unidad_header_font
            unidad_cell_excel.alignment = left_alignment
            current_row += 1
            for col_num_excel, header_title_excel in enumerate(data_headers, 1):
                cell_excel = sheet.cell(row=current_row, column=col_num_excel, value=header_title_excel)
                cell_excel.font = header_font
                cell_excel.alignment = centered_alignment
                cell_excel.fill = header_fill
                cell_excel.border = header_border
            current_row += 1
            nro_item_unidad_excel = 0
            for detalle_item_excel in detalles_de_esta_unidad_excel:
                nro_general_item_excel += 1
                nro_item_unidad_excel +=1

                row_data_excel = [
                    nro_item_unidad_excel, 
                    getattr(detalle_item_excel, 'item_externo', ''),
                    getattr(detalle_item_excel, 'ci_externo', ''),
                    getattr(detalle_item_excel, 'nombre_completo_externo', ''),
                    getattr(detalle_item_excel, 'cargo_externo', ''),
                    safe_decimal(planilla.dias_habiles, default=Decimal('0.00')),
                    safe_decimal(getattr(detalle_item_excel, 'faltas', None)),
                    safe_decimal(getattr(detalle_item_excel, 'vacacion', None)),
                    safe_decimal(getattr(detalle_item_excel, 'viajes', None)),
                    safe_decimal(getattr(detalle_item_excel, 'bajas_medicas', None)),
                    safe_decimal(getattr(detalle_item_excel, 'pcgh', None)),
                    safe_decimal(getattr(detalle_item_excel, 'psgh', None)),
                    safe_decimal(getattr(detalle_item_excel, 'perm_excep', None)),
                    safe_decimal(getattr(detalle_item_excel, 'asuetos', None)),
                    safe_decimal(getattr(detalle_item_excel, 'pcgh_embar_enf_base', None)),
                    safe_decimal(getattr(detalle_item_excel, 'dias_no_pagados', None)),
                    safe_decimal(getattr(detalle_item_excel, 'dias_pagados', None)),
                    safe_decimal(getattr(detalle_item_excel, 'total_ganado', None)),
                    safe_decimal(getattr(detalle_item_excel, 'descuentos', None)),
                    safe_decimal(getattr(detalle_item_excel, 'liquido_pagable', None))
                    
                ]
                for col_idx_excel, cell_value_excel in enumerate(row_data_excel, 1):
                    cell_obj_excel = sheet.cell(row=current_row, column=col_idx_excel, value=cell_value_excel)
                    header_name_actual_excel = data_headers[col_idx_excel-1]
                    cell_obj_excel.font = value_font

                    if header_name_actual_excel in ['Nro.', 'Item', 'CI']:
                        cell_obj_excel.alignment = centered_alignment
                        if header_name_actual_excel == 'Item' and (isinstance(cell_value_excel, int) or (isinstance(cell_value_excel, str) and cell_value_excel.isdigit())):
                            cell_obj_excel.number_format = '0'
                    elif header_name_actual_excel in ['Nombre Completo', 'Cargo']:
                        cell_obj_excel.alignment = left_alignment
                    else: 
                        cell_obj_excel.alignment = right_alignment
                        if isinstance(cell_value_excel, Decimal):
                            # Mostrar 2 decimales para la mayoría de los campos Decimal
                            is_monetary_field = header_name_actual_excel in ['Total Ganado', 'Desc.', 'Líquido Pagable']
                            if cell_value_excel == cell_value_excel.to_integral_value() and not is_monetary_field and header_name_actual_excel != 'Días Háb.':
                                 cell_obj_excel.number_format = integer_format
                            else:
                                 cell_obj_excel.number_format = decimal_format
                        elif isinstance(cell_value_excel, int): 
                             cell_obj_excel.number_format = integer_format
                current_row += 1
            current_row += 1 
        # --- Ajustar Ancho de Columnas ---
        anchos_especificos = {
            'Nro.': 6, 'Item': 8, 'CI': 11, 'Nombre Completo': 25, 'Cargo': 30,
            'Días Háb.': 8, 'Faltas': 7, 'Vacac.': 7, 'Viajes': 7, 'B.Médicas': 7, 'PCGH': 7,
            'PSGH': 7, 'P.Excep': 7, 'Asuetos': 7, 'PCGH Emb/Enf': 7,
            'D.No Pag.': 8, 'D. Pag.': 8,
            'Total Ganado': 10, 'Desc.': 9, 'Líquido Pagable': 11,
            
        }
        ancho_default_excel = 9
        for col_idx_aw_excel, header_name_aw_excel in enumerate(data_headers, 1):
            column_letter_excel = get_column_letter(col_idx_aw_excel)
            adjusted_width_excel = anchos_especificos.get(header_name_aw_excel, ancho_default_excel)
            sheet.column_dimensions[column_letter_excel].width = adjusted_width_excel
        logger.debug("Ancho de columnas ajustado para Excel.")

    except Exception as e_xlsx_build:
        logger.error(f"Error crítico construyendo el archivo XLSX para Planilla {planilla_id}: {e_xlsx_build}", exc_info=True)
        messages.error(request, f"Error interno al generar el archivo Excel: {e_xlsx_build}")
        return redirect('ver_detalles_bono_te', planilla_id=planilla_id)

    # --- Crear la Respuesta HTTP ---
    try:
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        tipo_planilla_safe_excel = "".join(c if c.isalnum() else '_' for c in planilla.get_tipo_display())
        filename_base_excel = f"BonoTE_{tipo_planilla_safe_excel}_{planilla.mes}_{planilla.anio}"
        
        filter_suffix_parts_excel = []
        if selected_secretaria_id: filter_suffix_parts_excel.append(f"Sec{selected_secretaria_id}")
        if selected_unidad_id: filter_suffix_parts_excel.append(f"Uni{selected_unidad_id}")
        if search_term: 
            safe_search_term = "".join(c if c.isalnum() else '_' for c in search_term[:15])
            filter_suffix_parts_excel.append(f"Q_{safe_search_term}")
        
        filter_suffix_excel = "_".join(filter_suffix_parts_excel)
        if filter_suffix_excel:
            filename_excel = f"{filename_base_excel}_Filtro_{filter_suffix_excel}.xlsx"
        else:
            filename_excel = f"{filename_base_excel}.xlsx"

        response['Content-Disposition'] = f'attachment; filename="{filename_excel}"'
        workbook.save(response)
        logger.info(f"Archivo XLSX '{filename_excel}' generado y enviado para Planilla {planilla.id}. Total detalles procesados: {total_detalles_a_exportar}")
        return response
    except Exception as e_resp_excel:
        logger.error(f"Error creando o enviando HttpResponse para XLSX (Planilla {planilla_id}): {e_resp_excel}", exc_info=True)
        messages.error(request, f"Error final al preparar la descarga del archivo: {e_resp_excel}")
        return redirect('ver_detalles_bono_te', planilla_id=planilla_id)
    

#-------------------------------------
@login_required
# @permission_required('planilla.view_planilla', raise_exception=True) # O un permiso más específico
def export_detalles_bonote_pdf(request, planilla_id):
    logger.info(f"Solicitud de exportación PDF para Detalles Bono TE de Planilla ID: {planilla_id}")
    processed_data = get_processed_planilla_details(request, planilla_id, return_all_for_export=True)
    planilla_cabecera = processed_data.get('planilla')
    detalles_agrupados = processed_data.get('detalles_agrupados_por_unidad_export')
    orden_unidades = processed_data.get('orden_unidades_export')
    error_message = processed_data.get('error_message')

    if error_message and not planilla_cabecera:
        logger.error(f"Error al obtener datos para exportar PDF Detalles Bono TE (Planilla ID {planilla_id}): {error_message}")
        messages.error(request, f"No se pudo generar el PDF: {error_message or 'Planilla no encontrada.'}")
        return redirect('lista_planillas') 

    if not planilla_cabecera: 
        messages.error(request, "Planilla no encontrada.")
        return redirect('lista_planillas')

    if not detalles_agrupados or not orden_unidades:
        logger.warning(f"Planilla Bono TE ID {planilla_id} no tiene detalles (o no se pudieron agrupar) para exportar a PDF.")
        messages.warning(request, "La planilla no tiene detalles (o no coinciden con filtros) para exportar.")
        return redirect('ver_detalles_bono_te', planilla_id=planilla_id)


    # 2. Definir las columnas para el PDF de Detalles Bono TE
    column_definitions_bonote = [
        ('Nro.', 0.3*inch, 'nro_item'), 
        ('Item', 0.4*inch, 'item_externo'),
        ('CI', 0.6*inch, 'ci_externo'),
        ('Nombre Completo', 1.5*inch, 'nombre_completo_externo'),
        ('Cargo', 1.6*inch, 'cargo_externo'),
        ('D.Háb.', 0.35*inch, 'dias_habiles_planilla'), 
        ('Faltas', 0.35*inch, 'faltas'),
        ('Vacac.', 0.35*inch, 'vacacion'),
        ('Viajes', 0.35*inch, 'viajes'),
        ('B.Méd', 0.35*inch, 'bajas_medicas'),
        ('PCGH', 0.35*inch, 'pcgh'),
        ('PSGH', 0.35*inch, 'psgh'),
        ('P.Exc', 0.35*inch, 'perm_excep'),
        ('Asuetos', 0.35*inch, 'asuetos'),
        ('PCGH.E', 0.35*inch, 'pcgh_embar_enf_base'),
        ('D.NoPag', 0.35*inch, 'dias_no_pagados'),
        ('D.Pag', 0.35*inch, 'dias_pagados'),
        ('Total.G', 0.5*inch, 'total_ganado'),
        ('Desc.', 0.5*inch, 'descuentos'),
        ('Líq.Pag.', 0.5*inch, 'liquido_pagable'),
    ]
    # 3. Preparar respuesta HTTP y buffer
    response = HttpResponse(content_type='application/pdf')
    tipo_planilla_str = "".join(c if c.isalnum() else '_' for c in planilla_cabecera.get_tipo_display())
    filename = f"Detalles_BonoTE_{tipo_planilla_str}_{planilla_cabecera.mes}_{planilla_cabecera.anio}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    buffer = io.BytesIO()

    # 4. Llamar a la función de utilidad para generar el PDF
    try:
        generar_pdf_bonote_detalles(
            output_buffer=buffer,
            planilla_cabecera_bonote=planilla_cabecera,
            detalles_agrupados_bonote=detalles_agrupados,
            orden_unidades_bonote=orden_unidades,
            column_definitions_bonote=column_definitions_bonote
        )
    except Exception as e_gen_pdf:
        logger.error(f"Vista: Falla al llamar a generar_pdf_bonote_detalles para Planilla ID {planilla_id}: {e_gen_pdf}", exc_info=True)
        messages.error(request, f"Ocurrió un error crítico al generar el documento PDF: {e_gen_pdf}")
        return redirect('ver_detalles_bono_te', planilla_id=planilla_id)

    # 5. Escribir el PDF en la respuesta
    pdf_data = buffer.getvalue()
    buffer.close()
    response.write(pdf_data)

    logger.info(f"Vista: PDF de Detalles Bono TE generado y enviado para Planilla ID {planilla_id}.")
    return response



#------------------------------------------
@login_required
# @permission_required('planilla.view_planilla', raise_exception=True) # O el permiso adecuado
def export_lista_planillas_pdf(request):
    logger.info("Solicitud de exportación PDF para la lista de planillas.")
    planillas_a_exportar = Planilla.objects.all().order_by('-anio', '-mes', 'tipo')
    column_definitions_lista = [
        ("Nro.", 0.5*inch, 'nro_item_lista', 'center'),
        ("Mes/Año", 1.0*inch, lambda p: f"{p.mes}/{p.anio}", 'center'),
        ("Tipo", 1.5*inch, 'get_tipo_display', 'left'), 
        ("Estado", 1.0*inch, 'get_estado_display', 'left'), 
        ("Días Háb.", 0.8*inch, 'dias_habiles', 'right'),
        ("Elaborado por", 1.2*inch, 'usuario_elaboracion', 'left'),
        ("Fecha Elab.", 1.0*inch, 'fecha_elaboracion', 'center'),
    ]
    # Ancho total aprox: 0.5+1.0+1.5+1.0+0.8+1.2+1.0 = 7.0 pulgadas. Cabe bien en portrait letter (8.5 - márgenes).

    response = HttpResponse(content_type='application/pdf')
    filename = "lista_planillas_bonote.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    buffer = io.BytesIO()

    try:
        generar_pdf_lista_planillas(
            output_buffer=buffer,
            planillas_qs=planillas_a_exportar,
            column_definitions_lista=column_definitions_lista,
            titulo_reporte="LISTA DE PLANILLAS DE BONO DE TÉ" 
        )
    except Exception as e_gen_pdf:
        logger.error(f"Vista: Falla al llamar a generar_pdf_lista_planillas: {e_gen_pdf}", exc_info=True)
        messages.error(request, f"Ocurrió un error crítico al generar el documento PDF: {e_gen_pdf}")
        return redirect('lista_planillas') 

    pdf_data = buffer.getvalue()
    buffer.close()
    response.write(pdf_data)

    logger.info(f"PDF de lista de planillas generado y enviado.")
    return response